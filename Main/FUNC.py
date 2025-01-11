import shelve
import openpyxl
import algorithms
import tkinter as tk
from CTkTable import *
import csv
import customtkinter as Ct
import Knockout_format

data_file='INDOOR.db'

def autoinsert(c,adno,records):
    try:
        std=shelve.open('student_database.db',flag='r')
    except:
        f=open(c,'r')
        dict={}
        reader=csv.reader(f)
        d=shelve.open(data_file)
        std=shelve.open('student_database.db')
        for row in reader:
            dict[row[0]]=list(row[1:])
        std.update(dict)
        f.close()
    d=shelve.open(data_file)
    if adno.strip() in list(std.keys()):
        if records=='':
            tk.messagebox.showerror("Error",'Update Records') 
        else:
            for ad in std.keys():
                        
                if ad==adno:
                    key=ad
                    d[key]=std[ad]+[records]
                    tk.messagebox.showinfo('SUCESS','Student entered sucessfully')
                    break
            else:
                tk.messagebox.showerror("Error",'Student does not exist')
    d.close()


'''def maninsert(adno,name,clas,records): 
    
    row={} 
    d=shelve.open(data_file)
    if adno !='':
        if name !='':          
            if clas!='':             
                if records=='':
                    records='N'
                    row[adno.strip()]=list(row[name,clas,records])
                    d.update(row)
                if records!='':
                    row[adno.strip()]=list(row[name,clas,records])
                    d.update(row)
                

            tk.messagebox.showerror("Error",' Enter Class and Division')
        tk.messagebox.showerror("Error",' Enter Name')
    tk.messagebox.showerror("Error",' Enter Admission number')'''



def disp_all(page,page31):
    f=shelve.open(data_file)
    l=[['Ad_No','NAME','CLASS/DIVISION',"RECORDS"]]
    for i in f:
        l.append([i]+f[i])

    table = CTkTable(master=page,  column=4, values=l,width=1,anchor='w')
    table.pack(expand=True, fill="both", padx=20, pady=20)
    f.close()

    Print_button = Ct.CTkButton(page31,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Print(l))
    Print_button.place(relx=0.45, rely=0.9)

def Print(text):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    a=1
    b=0
    for i in text:
        b=b+1
        a=1
        for r in range(0,len(i)):
            sheet.cell(row=b,column=a).value=i[r]
            a=a+1
    wb.save('Participants-list.xlsx')
    tk.messagebox.showinfo("Sucess",' Excel file created')
    
def search(page,adno):
    f=shelve.open(data_file)
    l=[]
    if adno.strip() in list(f.keys()):
        for ad in f.keys():           
            if ad==adno:
                l=[['Ad_No','NAME','CLASS/DIVISION',"RECORDS"]]
    
                l.append([ad]+f[ad])

                table = CTkTable(master=page,  column=4, values=l)
                table.pack(expand=True, fill="both", padx=20, pady=20)
    else:
        tk.messagebox.showerror("Error",'Enter a VALID Admission number')
    f.close()

def delete(adno):
    f=shelve.open(data_file)
    for ad in list(f.keys()):
        if adno.strip()==ad:                              
            del f[ad] 
            tk.messagebox.showinfo('SUCESS','Student removed sucessfully')
            break
    else:
        tk.messagebox.showerror("Error",' Enter a VALID Admission number')
    
    f.close() 

'''
def modify(adno,name,clas,records,a,n,c,r):
    d=shelve.open(data_file)
    
    if adno.strip() in list(d.keys()):
        for ad in d.keys():           
            if ad==adno:
                a.configure(state="disabled")
                n.configure(state="normal")
                c.configure(state="normal")
                r.configure(state="normal")
                tk.messagebox.showinfo('INFO','Modify student details')
                
            if ad==adno:
                d[ad][0]==name
                d[ad][1]==clas
                d[ad][2]==records
                break
        tk.messagebox.showinfo('SUCESS','Student details modified sucessfully')
    
    else:
        tk.messagebox.showerror("Error",' Enter a VALID Admission number')
    '''    

def modify(adno,name,clas,records):
    f=shelve.open(data_file)
    
    if adno.strip() in list(f.keys()):
        for ad in f.keys():           
            if ad==adno:
               
                
                f[ad][0]==name
                f[ad][1]==clas
                f[ad][2]==records
                break
        tk.messagebox.showinfo('SUCESS','Student details modified sucessfully')
    
    else:
        tk.messagebox.showerror("Error",' Enter a VALID Admission number')



def create_fixtures(page):
    f=shelve.open(data_file)
    First_Round=algorithms.create_fixtures(data_file)
    
    button_list=[]
    
    page41=Ct.CTkScrollableFrame(page,fg_color='#071952',bg_color='#687EFF',
                    width=1100,height=650,border_width=3,border_color='#687EFF')

    page41.place(relx=0.18,rely=0.02)
    button3 = Ct.CTkButton(page,bg_color=('transparent'),text='SUBMIT',
                    font=("Plasma",20,'bold' ),
                    fg_color=('transparent')  ,width=150, height=75 ,
                    command=lambda:edit_result1(button_list)) 
            
    if len(f)>16:

        for i,j in enumerate(First_Round):

            frame1=Ct.CTkFrame(page41,fg_color='#071952',bg_color='#687EFF',
                       width=875,height=290,
                       border_width=3,border_color='#687EFF')                       
        
            frame1.grid(row=i+1,column=0)

            label1=Ct.CTkLabel( frame1,text=" MATCH NO "+ str(i+1),
                            font=("Plasma",13.5,'bold'),width=100,height=50)
            label1.grid(row=0,column=0)

            button1=tk.Button(frame1,text=j[0][:-1]
                              ,width=45,height=2,
                             bg='blue', font=("Plasma",12,'bold'))
            button2=tk.Button(frame1,text=j[1][:-1]
                              ,width=45,height=2,
                              bg='blue', font=("Plasma",12,'bold'))

            button1.grid(row=0,column=1)
            button1.configure(command=lambda a=button1,b=button2:Select(a,b))

            label2=Ct.CTkLabel(frame1,text= "VS",
                            font=("Plasma",15,'bold'),width=50,height=50)
            label2.grid(row=0,column=2)

            button2.grid(row=0,column=3)
            button2.configure(command=lambda a=button2,b=button1:Select(a,b))

            button_list.extend([button1,button2])

            button3 = Ct.CTkButton(page,bg_color=('transparent'),text='SUBMIT',
                    font=("Plasma",20,'bold' ),
                    fg_color=('transparent')  ,width=150, height=75 ,
                    command=lambda:edit_result1(button_list)) 
            
            button3.place(relx=0.05, rely=0.15)
            
            Print_button = Ct.CTkButton(page,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Print1(First_Round))
            Print_button.place(relx=0.06, rely=0.3)

    else:
        page41.forget()
        button3.forget()
        page41=Ct.CTkFrame(page,fg_color='#071952',bg_color='#687EFF',
                    width=1350,height=675,border_width=3,border_color='#687EFF')

        page41.place(relx=0.01,rely=0.02)
        
        Knockout_format.match_format(page41)

        
    def edit_result1(button_list):
        for i in button_list:
            if i.cget('bg')=='blue':
                tk.messagebox.showerror("Error",'Update All The Matches ')    
                break
        else:
            deselected_button=[a.cget('text') for a in button_list if a.cget('bg')=='red']
            
            for i in deselected_button:
                delete1(i[:6])
            tk.messagebox.showinfo('SUCESS','Student moved to the next Round')
            create_fixtures(page)

def Print1(text):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    a=1
    b=0
    l=[['Ad_No','NAME','','Ad_No','NAME']]

    for i in text:  
        l.append(i[0][:2]+['VS']+i[1][:2])

    for i in l:   
        b=b+1
        a=1                      
        for r in range(0,len(i)):
            sheet.cell(row=b,column=a).value=i[r]
            a=a+1
    wb.save('fixtures-list.xlsx')
    tk.messagebox.showinfo("Sucess",' Excel file created')
              
def Select(a,b):
    if a.cget('bg')!='green':  
        a.configure(bg='green')
        b.configure(bg='red')
    else:
        a.configure(bg='red')
        b.configure(bg='green')
    
def delete1(adno):
    
    f=shelve.open(data_file)
    for ad in list(f.keys()):
        if adno.strip()==ad:                              
            del f[ad]  
            break
    else:
        tk.messagebox.showerror("Error",' Enter a VALID Admission number')
    
    f.close() 


'''
def Next_Round(button_list):
    for i in button_list:
        if i.cget('bg')=='blue':
            tk.messagebox.showerror("Error",'Update All The Matches ')    
            break
    else:
        deselected_button=[a.cget('text') for a in button_list if a.cget('bg')=='red']
        
        for i in deselected_button:
            delete1(i[:6])
    

'''








'''
    l=[["MATCH.NO",'AD.NO','PLAYER 1',' ' ,'AD.NO','PLAYER 2']]
    a=0

    for i in First_Round:
        a=a+1
        l.append([a]+[i[0][0]]+[i[0][1]]+['VS']+[i[1][0]]+[i[1][1]])

    table = CTkTable(page,  column=6, values=l)
    table.pack(expand=True, fill="both", padx=20, pady=20)
   
    f.close()'''