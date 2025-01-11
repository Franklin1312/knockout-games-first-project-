import customtkinter as Ct
import tkinter as tk
import os
import shelve
import csv
from CTkTable import *
from PIL import Image
import FUNC
import Knockout_format
import openpyxl 

Ct.set_appearance_mode('dark')

root=Ct.CTk()   

root.geometry('1366x720')
root.title("Made by Franklin Babu and Dhishaj C")

path=os.path.dirname(os.path.abspath(__name__))
path=path.replace('\\','/')+'/'



#WINDOW 1
    
page1=Ct.CTkFrame(root)
page1.pack(expand=True,fill=tk.BOTH)

bg = Ct.CTkImage( light_image=Image.open('1.jpg'),size=(1366,720) )
pic1 = Ct.CTkLabel(page1, image = bg,text='') 
pic1.place(relx=0, rely=0)


label1= Ct.CTkLabel(page1,text="GAMES MANAGEMENT",
                    width=800,height=100,corner_radius=0,bg_color=('white'),
                    fg_color=('transparent'),text_color='blue',
                    font=("Broadway",100,'bold' )  )
label1.place(relx=0.5, rely=0.3, anchor=Ct.CENTER)

entry1 = Ct.CTkEntry(page1, placeholder_text="Enter Password",
                    corner_radius=0,fg_color=('transparent'),
                    width=200,height=40,border_width=3,border_color='white'
                    )
                    
entry1.place(relx=0.43, rely=0.5)

login1 = Ct.CTkButton(page1, text="Login",font=("Stylus",25,'bold' ),corner_radius=0,
                       fg_color=('transparent') ,bg_color=('transparent'),
                       border_width=3,border_color='white',
                       command=lambda:login() )
login1.place(relx=0.45, rely=0.7)

Indoor_button = tk.Button(root,bg=('blue'),text='INDOOR GAMES',
                font=("Plasma",10,'bold' ),borderwidth=3,
                fg=('black')  ,width=19, height=2  
                 )
Indoor_button.place(relx=0.52, rely=0.6)

Outdoor_button = tk.Button(root,bg=('blue'),text='OUTDOOR GAMES',
            font=("Plasma",10,'bold' ),borderwidth=3,
            fg=('black')  ,width=19, height=2  
             )
Outdoor_button.place(relx=0.38, rely=0.6)

Indoor_button.configure(command=lambda a=Indoor_button,b=Outdoor_button:Select(a,b))
Outdoor_button.configure(command=lambda a=Outdoor_button,b=Indoor_button:Select(a,b))

def Select(a,b):
        if a.cget('bg')!='green':  
            a.configure(bg='green')
            b.configure(bg='red')
        else:
            a.configure(bg='red')
            b.configure(bg='green')

def login():
    global entry1,page2,page1,Outdoor_button,Indoor_button
    if entry1.get()=='nmcs123': 
    
        for i in range (2):
            if Indoor_button.cget('bg')=='green':
                page1.destroy()
                page_2()
            else:
                page1.destroy()
                first_page()
                

    else:
        tk.messagebox.showerror("Error",'Wrong Password')
        
    

#WINDOW 2
page2=Ct.CTkFrame(root,fg_color='white')

page21=Ct.CTkFrame(page2,fg_color='#071952',bg_color='#687EFF',
                       width=500,height=400,border_width=3,border_color='#687EFF')

page22=Ct.CTkFrame(page2,fg_color='#DA0C81',bg_color='#DA0C81',
                       width=500,height=400)
    

def page_2():
    global root,page2,page22,page21
    
    
    page2.pack(padx=0,pady=0,expand=True,fill=tk.BOTH)
    page2.tkraise()
    

    bg2 = Ct.CTkImage( light_image=Image.open('2.jpg'),size=(1366,720) )
    pic2 = Ct.CTkLabel(page2, image = bg2,text='') 
    pic2.place(relx=0, rely=0)

    page21=Ct.CTkFrame(page2,fg_color='#071952',bg_color='#687EFF',
                       width=500,height=400,border_width=3,border_color='#687EFF')
    page21.place(relx=0.115,rely=0.3)
    page22=Ct.CTkFrame(page2,fg_color='#DA0C81',bg_color='#DA0C81',
                       width=500,height=400)
    page22.place(relx=0.53,rely=0.15)
    
    label2=Ct.CTkLabel(page2,text="Student Management",
                    width=500,height=100,corner_radius=0,bg_color=('transparent'),
                    fg_color=('grey'),text_color='blue',
                    font=("Plasma",40,'bold' )  )
    
    label2.place(relx=0.115, rely=0.15)

    Student_bg3 = Ct.CTkImage( light_image=Image.open('3.png'),size=(250,250) )
    button1 = Ct.CTkButton(page21,image=Student_bg3,text='',
                       border_width=3,border_color='#687EFF',
                       width=0,height=0,     
                       command=lambda:stdm() )
    
    button1.place(relx=0.23, rely=0.20)

    
    
    label2=Ct.CTkLabel(page2,text="GAME FIXTURES",
                    width=500,height=100,corner_radius=0,bg_color=('transparent'),
                    fg_color=('grey'),text_color='blue',
                    font=("Plasma",40,'bold' )  )
    
    label2.place(relx=0.53, rely=0.725)

    Game_Fix_bg4 = Ct.CTkImage( light_image=Image.open('4.png') ,size=(250,250) )
    button2 = Ct.CTkButton(page22,image=Game_Fix_bg4,text='',
                           width=250,height=250,
                       border_width=3,border_color='#687EFF',       
                       command=lambda:gamef() )
    
    button2.place(relx=0.2, rely=0.15)


#WINDOW 3
def stdm():
    global page2,root,page21
    page2.pack_forget()
    page_3()


page3=Ct.CTkFrame(root,fg_color='white')

page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                       width=800,height=500,
                       border_width=3,border_color='#687EFF')

page311=Ct.CTkFrame(page31,fg_color='#071952',bg_color='transparent',
                    width=700,height=300,
                    border_width=3,border_color='black')

def page_3():
    global page21,page2,page3

    page3.pack(padx=0,pady=0,expand=True,fill=tk.BOTH)
    page3.tkraise()

    bg5 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic3 = Ct.CTkLabel(page3, image = bg5,text='') 
    pic3.place(relx=0, rely=0)

    button1 = Ct.CTkButton(page3,text='DISPLAY',text_color='gold',
                       bg_color=('#687EFF'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),
                       width=100,height=50,     
                       command=lambda:display1() )

    button1.place(relx=0.0, rely=0.35)


    button2 = Ct.CTkButton(page3,text='INSERT',bg_color=('#687EFF'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=100,height=50,     
                       command=lambda:Insert1() )
    
    button2.place(relx=0.0, rely=0.45)


    button3 = Ct.CTkButton(page3,text='SEARCH',bg_color=('#687EFF'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=100,height=50,     
                       command=lambda:Search1() )
    
    button3.place(relx=0.0, rely=0.55)

    button4 = Ct.CTkButton(page3,text='MODIFY',bg_color=('#687EFF'),
                           border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=100,height=50,     
                       command=lambda:Modify1() )
    
    button4.place(relx=0, rely=0.65)


    button5 = Ct.CTkButton(page3,text='DELETE',bg_color=('#687EFF'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=100,height=50,     
                       command=lambda:delete1() )
    
    button5.place(relx=0, rely=0.75)


    Go_Back_bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,30))
    button6 = Ct.CTkButton(page3,image= Go_Back_bg6,bg_color=('transparent'),text='',
                       fg_color=('transparent')  ,width=100, height=30  ,
                       command=lambda:GoBack1() )
    
    
    button6.place(relx=0.0, rely=0.0)

    page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                       width=800,height=500,
                       border_width=3,border_color='#687EFF')


    page31.place(relx=0.2,rely=0.15)

def Insert1():
    global page31,page3
    page31.pack_forget()
    page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                    width=800,height=500,
                    border_width=3,border_color='#687EFF')


    page31.place(relx=0.2,rely=0.15)

    entry1 = Ct.CTkEntry(page31, placeholder_text="Enter ",
            corner_radius=0,fg_color=('transparent'),
                width=475,height=40,bg_color=('transparent')
                )
                
    entry1.place(relx=0.37, rely=0.2)

    label1= Ct.CTkLabel(page31,text="AD NO                                   :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label1.place(relx=0.009, rely=0.21)

    entry4 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
    entry4.place(relx=0.37, rely=0.31)

    label4= Ct.CTkLabel(page31,text="RECORD                               :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label4.place(relx=0.009, rely=0.31)


    button1 = Ct.CTkButton(page31,bg_color=('transparent'),text='SUBMIT',
                    fg_color=('transparent')  ,width=100, height=20  ,
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:(submitinsert(entry1,entry4)) )


    button1.place(relx=0.43, rely=0.9)

    def submitinsert(entry1,entry4):
        
        FUNC.autoinsert('classes 1-12.csv',entry1.get(),entry4.get())

''' def manual():
        page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                    width=800,height=500,
                    border_width=3,border_color='#687EFF')


        page31.place(relx=0.2,rely=0.15)

        entry1 = Ct.CTkEntry(page31, placeholder_text="Enter ",
            corner_radius=0,fg_color=('transparent'),
                width=475,height=40,bg_color=('transparent')
                )
                
        entry1.place(relx=0.37, rely=0.2)

        label1= Ct.CTkLabel(page31,text="AD NO                                   :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
        label1.place(relx=0.009, rely=0.21)

        entry2 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
        entry2.place(relx=0.37, rely=0.36)

        label2= Ct.CTkLabel(page31,text="NAME                                    :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
        label2.place(relx=0.009, rely=0.37)

        entry3 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
        entry3.place(relx=0.37, rely=0.52)

        label3= Ct.CTkLabel(page31,text="CLASS/DIV                            :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
        label3.place(relx=0.009, rely=0.53)


        entry4 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
        entry4.place(relx=0.37, rely=0.66)

        label4= Ct.CTkLabel(page31,text="RECORD                               :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
        label4.place(relx=0.009, rely=0.67)

        entry2.configure(state="disabled")
        entry3.configure(state="disabled")

        button1 = Ct.CTkButton(page31,bg_color=('transparent'),text='SUBMIT',
                    fg_color=('transparent')  ,width=100, height=20  ,
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:(submitinsert1(entry1,entry2,entry3,entry4)) )


        button1.place(relx=0.43, rely=0.9)

        def submitinsert1(entry1,entry2,entry3,entry4):
        
            FUNC.maninsert(entry1.get(),entry2.get(),entry3.get(),entry4.get())'''

def Modify1():
    global page31,page3
    page31.pack_forget()
    page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                    width=800,height=500,
                    border_width=3,border_color='#687EFF')


    page31.place(relx=0.2,rely=0.15)
    entry1 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40,bg_color=('transparent')
                )
                
    entry1.place(relx=0.37, rely=0.2)

    label1= Ct.CTkLabel(page31,text="AD NO                                   :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label1.place(relx=0.009, rely=0.21)

    entry2 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
    entry2.place(relx=0.37, rely=0.36)

    label2= Ct.CTkLabel(page31,text="NAME                                    :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label2.place(relx=0.009, rely=0.37)

    entry3 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
    entry3.place(relx=0.37, rely=0.52)

    label3= Ct.CTkLabel(page31,text="CLASS/DIV                            :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label3.place(relx=0.009, rely=0.53)


    entry4 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40
                )
                
    entry4.place(relx=0.37, rely=0.66)

    label4= Ct.CTkLabel(page31,text="RECORD                               :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label4.place(relx=0.009, rely=0.67)

    


    button1 = Ct.CTkButton(page31,bg_color=('transparent'),text='SUBMIT',
                    fg_color=('transparent')  ,width=100, height=20  ,
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:modify1(entry1,entry2,entry3,entry4) )

    button1.place(relx=0.43, rely=0.9)

    def modify1(entry1,entry2,entry3,entry4):
        FUNC.modify(entry1.get(),entry2.get(),entry3.get(),entry4.get())

def Search1():
    global page31,page3
    page31.pack_forget()
    page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                    width=800,height=500,
                    border_width=3,border_color='#687EFF')


    page31.place(relx=0.2,rely=0.15)

    entry1 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40,bg_color=('transparent')
                )
                
    entry1.place(relx=0.37, rely=0.1)

    label1= Ct.CTkLabel(page31,text="AD NO                                   :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label1.place(relx=0.009, rely=0.11)


    button1 = Ct.CTkButton(page31,bg_color=('transparent'),text='SUBMIT',
                    fg_color=('transparent')  ,width=100, height=20  ,
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:search1(page311,entry1) )


    button1.place(relx=0.43, rely=0.25)


    page311=Ct.CTkFrame(page31,fg_color='#83A2FF',bg_color='transparent',
                    width=700,height=300,
                    border_width=3,border_color='black')

    page311.place(relx=0.08,rely=0.33)

    def search1(page311,entry1):
        FUNC.search(page311,entry1.get())

def delete1():
    global page31,page3
    page31.pack_forget()
    page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                    width=800,height=500,
                    border_width=3,border_color='#687EFF')


    page31.place(relx=0.2,rely=0.15)

    entry1 = Ct.CTkEntry(page31, placeholder_text="Enter ",
                corner_radius=0,fg_color=('transparent'),
                width=475,height=40,bg_color=('transparent')
                )
                
    entry1.place(relx=0.37, rely=0.1)

    label1= Ct.CTkLabel(page31,text="AD NO                                   :",
                corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    label1.place(relx=0.009, rely=0.11)


    button1 = Ct.CTkButton(page31,bg_color=('transparent'),text='SUBMIT',
                    fg_color=('transparent')  ,width=100, height=20  ,
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:delete1(entry1) )

    button1.place(relx=0.43, rely=0.25)

    page311=Ct.CTkFrame(page31,fg_color='#071952',bg_color='transparent',
                    width=700,height=300,
                    border_width=3,border_color='black')

    page311.place(relx=0.08,rely=0.33)

    def delete1(entry1):
        FUNC.delete(entry1.get())

def display1():
    global page31,page3
    page31.pack_forget()
    page31=Ct.CTkFrame(page3,fg_color='#071952',bg_color='#687EFF',
                    width=800,height=500,
                    border_width=3,border_color='#687EFF')


    page31.place(relx=0.2,rely=0.15)

    button1 = Ct.CTkButton(page31,bg_color=('transparent'),text='SHOW DATA',
                    fg_color=('transparent')  ,width=200, height=40  ,
                    font=("Stylus",35,'bold' ) ,text_color='White',
                    command=lambda:disp_all(page311,page31) )


    button1.place(relx=0.35, rely=0.05)

    page311=Ct.CTkScrollableFrame(page31,fg_color='#83A2FF',bg_color='transparent',
                    width=750,height=350,
                    border_width=5,border_color='black')

    page311.place(relx=0.02,rely=0.15)

    def disp_all(page311,page31):
        FUNC.disp_all(page311,page31)


def GoBack1():
    global page3,root,page21,page22,page2
    page3.pack_forget()
    page_2()


#WINDOW 4

def gamef():
    global page2,root,page21
    page2.pack_forget()
    page_4()

page4=Ct.CTkFrame(root,fg_color='white')


def page_4():
    global page22,page2,page4,page41
    
    page4=Ct.CTkFrame(root,fg_color='black',bg_color='transparent')
    page4.pack(padx=0,pady=0,expand=True,fill=tk.BOTH)
    page4.tkraise()

    bg7 = Ct.CTkImage( light_image=Image.open('7.jpg'),size=(1366,720) )
    pic7 = Ct.CTkLabel(page4, image = bg7,text='') 
    pic7.place(relx=0, rely=0)

    page41=Ct.CTkFrame(page4,fg_color='#071952',bg_color='#687EFF',
                    width=1250,height=675,border_width=3,border_color='#687EFF')

    page41.place(relx=0.085,rely=0.02)
    
    button1 = Ct.CTkButton(page4,bg_color=('#071952'),text='START',
                       font=("Plasma",20,'bold' ),
                       fg_color=('#071952')  ,width=175, height=75  ,
                       command=lambda:create_match() )
    
    button1.place(relx=0.47, rely=0.865)

    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    button2 = Ct.CTkButton(page4,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:GoBack_2() )
    
    button2.place(relx=0.0, rely=0.0)
    Knockout_format.match_format(page41)
    
    def create_match():
        global page22,page2,page4

        page4.pack_forget()
        
        page4=Ct.CTkFrame(root,fg_color='black',bg_color='transparent')
        page4.pack(padx=0,pady=0,expand=True,fill=tk.BOTH)
        page4.tkraise()

        bg7 = Ct.CTkImage( light_image=Image.open('7.jpg'),size=(1366,720) )
        pic7 = Ct.CTkLabel(page4, image = bg7,text='') 
        pic7.place(relx=0, rely=0)

        FUNC.create_fixtures(page4)

def GoBack_2():
    global page4,root,page21,page22,page2
    page4.pack_forget()
    page_2()


#-----------------------------------------------OUTDOOR GAMES-------------------------------------------
    
data_file='OUTDOOR.db'
var_list_kiddies=[]
var_list_subjunior=[]
var_list_junior=[]
var_list_senior=[]


def first_page():
    global page5,root,mainframe,mainframe2,mainframe_2,mainframe3,mainframe_3
    page5=Ct.CTkFrame(master=root,fg_color="#071952",
                    width=1366,height=720,border_width=3,border_color='black')
    page5.pack()

    bg71 = Ct.CTkImage( light_image=Image.open('7.jpg'),size=(1366,720) )
    pic71 = Ct.CTkLabel(page5, image = bg71,text='') 
    pic71.place(relx=0, rely=0)

    bg3 = Ct.CTkImage( light_image=Image.open('3.png'),size=(300,300) )
    StudentManagement = Ct.CTkButton(page5,image=bg3,bg_color=('transparent'),
                    text='',fg_color=('transparent')  ,
                    width=300, height=300  ,border_width=3,border_color='black',
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:Student_Management() )
    StudentManagement.place(relx=0.1, rely=0.4)

    stud1= Ct.CTkLabel(page5,text="STUDENT MANAGEMENT",
                corner_radius=0,bg_color=('transparent'),width=300,
                    fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    stud1.place(relx=0.1, rely=0.31)

        
    bg4 = Ct.CTkImage( light_image=Image.open('4.png') ,size=(300,300) )
    Matchfixtures = Ct.CTkButton(page5,image=bg4,bg_color=('transparent'),
                    text='',fg_color=('transparent')  ,
                    border_width=3,border_color='black',width=300, height=30,
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:Match_Fixtures() )
    Matchfixtures.place(relx=0.4, rely=0.4)

    stud2= Ct.CTkLabel(page5,text="MATCH FIXTURES",
                corner_radius=0,bg_color=('transparent'),width=300,
                    fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    stud2.place(relx=0.4, rely=0.31)

    bg5 = Ct.CTkImage( light_image=Image.open('8.jpg') ,size=(300,300) )    
    Result_generation = Ct.CTkButton(page5,image=bg5,bg_color=('transparent'),
                    text='',fg_color=('transparent'),
                    width=300, height=300  ,border_width=3,border_color='black',
                    font=("Stylus",20,'bold' ) ,
                    command=lambda:Result_Generation() )
    Result_generation .place(relx=0.7, rely=0.4)

    stud3= Ct.CTkLabel(page5,text="RESULT GENERATION",
                corner_radius=0,bg_color=('transparent'),width=300,
                    fg_color=('transparent'),text_color='gold',
                font=("Stylus",20,'bold' )  )
    stud3.place(relx=0.7, rely=0.31)

   
mainframe=Ct.CTkFrame(root,fg_color="#071952",
                          width=1366,height=720,border_color='black')

selected=[]
check_list=[]

def Student_Management():
    global page5,root,mainframe
    page5.pack_forget()
    
    mainframe.place(relx=0,rely=0)
    mainframe.tkraise()

    bg1 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic1 = Ct.CTkLabel(mainframe, image = bg1,text='') 
    pic1.place(relx=0, rely=0)

    mainframe1=Ct.CTkScrollableFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=800,height=490,
                    border_width=3,border_color='black')
    mainframe1.place(relx=0.218,rely=0.15)
    
    mainframe2=Ct.CTkFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=250,height=150,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.01,rely=0.08)
    
   
    enterAD_NO= Ct.CTkEntry(mainframe2, placeholder_text="Enter ",
        corner_radius=0,
            width=125,height=20
            )
                
    enterAD_NO.place(relx=0.45, rely=0.25)

    label1= Ct.CTkLabel(mainframe2,text="AD NO :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    label1.place(relx=0.1, rely=0.21)


    Submit1 = Ct.CTkButton(mainframe2,bg_color=('red'),text='SUBMIT',
                fg_color=('transparent')  ,width=5, height=2  ,
                font=("Stylus",20,'bold' ) ,
                command=lambda:Checking_Category("Classes 1-12.csv",enterAD_NO.get() ))
    Submit1.place(relx=0.35, rely=0.65)
    
    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback1 = Ct.CTkButton(mainframe,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:Goback1() )
    
    goback1.place(relx=0.0, rely=0.0)

def Goback1():
    global page5,root,mainframe
    mainframe.place_forget()
    first_page()

def kiddies(name_details,ad):                                                   #KIDDIES FUNCTIOS
    global page5,root,mainframe
    page5.pack_forget()

    category="Kiddies"
    bg1 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic3 = Ct.CTkLabel(mainframe, image = bg1,text='') 
    pic3.place(relx=0, rely=0)

    mainframe1=Ct.CTkScrollableFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=800,height=490,
                    border_width=3,border_color='black')
    mainframe1.place(relx=0.218,rely=0.15)
    
    mainframe.place(relx=0,rely=0)
    
    student_detailSFrame(name_details,ad,category,data_file)

    catagory_name=Ct.CTkLabel(mainframe,fg_color='lime',bg_color='transparent',
                    width=820,height=50,text='KIDDIES',
                    font=('Stylus',50,'bold'),text_color='white',                    
                    )
    catagory_name.place(relx=0.218,rely=0.05)

    mainframe2=Ct.CTkFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=250,height=150,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.01,rely=0.08)
   
    enterAD_NO= Ct.CTkEntry(mainframe2, placeholder_text="Enter ",
        corner_radius=0,
            width=125,height=20
            )
                
    enterAD_NO.place(relx=0.45, rely=0.25)

    label1= Ct.CTkLabel(mainframe2,text="AD NO :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    label1.place(relx=0.1, rely=0.21)


    Submit1 = Ct.CTkButton(mainframe2,bg_color=('red'),text='SUBMIT',
                fg_color=('transparent')  ,width=5, height=2  ,
                font=("Stylus",20,'bold' ) ,
                command=lambda:Checking_Category("Classes 1-12.csv",enterAD_NO.get() ))
    Submit1.place(relx=0.35, rely=0.65)
    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback1 = Ct.CTkButton(mainframe,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:Goback1() )
    
    goback1.place(relx=0.0, rely=0.0)

    list1=['50MTS','100MTS','STANDING BROAD JUMP','BALL THROW','4 x RELAY']
    check_list=[]

    for i in list1:
        v=tk.IntVar()
        b=Ct.CTkCheckBox(mainframe1,text=i,font=('Stylus',20),text_color='white',variable=v)
        
        b.configure(command=lambda b=b,name_details=name_details,ad=ad,
                    v=v: tracer(b,name_details,ad,category,v))               
        check_list.append(b)
        var_list_kiddies.append(v)

    for i,j in enumerate(check_list):
        n=3
        mainframe1.grid_rowconfigure(i//n,weight=2)
        mainframe1.grid_columnconfigure(i%n,weight=2)
        j.grid(row=i//n,
               column=i%n , sticky='W',padx=40,pady=20 )

def subjunior(name_details,ad):                                                 #SUB JUNIOR FUNCTIOS
    global page5,root,mainframe
    page5.pack_forget()
    
    category='Sub Junior'

    bg1 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic3 = Ct.CTkLabel(mainframe, image = bg1,text='') 
    pic3.place(relx=0, rely=0)

    mainframe1=Ct.CTkScrollableFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=800,height=490,
                    border_width=3,border_color='black')
    mainframe1.place(relx=0.218,rely=0.15)
    
    mainframe.place(relx=0,rely=0)
    
    student_detailSFrame(name_details,ad,category,data_file)

    catagory_name=Ct.CTkLabel(mainframe,fg_color='lime',bg_color='transparent',
                    width=820,height=50,text='SUB JUNIORS',
                    font=('Stylus',50,'bold'),text_color='white',                    
                    )
    catagory_name.place(relx=0.218,rely=0.05)
    
    mainframe2=Ct.CTkFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=250,height=150,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.01,rely=0.08)
    
   
    enterAD_NO= Ct.CTkEntry(mainframe2, placeholder_text="Enter ",
        corner_radius=0,
            width=125,height=20
            )
                
    enterAD_NO.place(relx=0.45, rely=0.25)

    label1= Ct.CTkLabel(mainframe2,text="AD NO :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    label1.place(relx=0.1, rely=0.21)

    Submit1 = Ct.CTkButton(mainframe2,bg_color=('red'),text='SUBMIT',
                fg_color=('transparent')  ,width=5, height=2  ,
                font=("Stylus",20,'bold' ) ,
                command=lambda:Checking_Category("Classes 1-12.csv",enterAD_NO.get() ))
    Submit1.place(relx=0.35, rely=0.65)
    
    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback1 = Ct.CTkButton(mainframe,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:Goback1() )
    
    goback1.place(relx=0.0, rely=0.0)

    list1=['100MTS','200MTS','400MTS','LONG JUMP','SHOT PUT','4 x RELAY']
    check_list=[]

    for i in list1:
        v=tk.IntVar()
        b=Ct.CTkCheckBox(mainframe1,text=i,font=('Stylus',20),text_color='white',variable=v)
        
        b.configure(command=lambda b=b,name_details=name_details,ad=ad,
                    v=v: tracer(b,name_details,ad,category,v))               
        check_list.append(b)
        var_list_subjunior.append(v)
    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
               column=i%n , sticky='W',pady=20 )
        mainframe1.grid_rowconfigure(i//n,weight=2)
        mainframe1.grid_columnconfigure(i%n,weight=2)
        
def junior(name_details,ad):                                                    #JUNIOR FUNCTION
    global page5,root,mainframe
    page5.pack_forget()

    category='Junior'

    bg1 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic3 = Ct.CTkLabel(mainframe, image = bg1,text='') 
    pic3.place(relx=0, rely=0)

    mainframe1=Ct.CTkScrollableFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=800,height=490,
                    border_width=3,border_color='black')
    mainframe1.place(relx=0.218,rely=0.15)

    mainframe.place(relx=0,rely=0)
    
    student_detailSFrame(name_details,ad,category,data_file)   

    catagory_name=Ct.CTkLabel(mainframe,fg_color='lime',bg_color='transparent',
                    width=820,height=50,text='JUNIORS',
                    font=('Stylus',50,'bold'),text_color='white',                    
                    )
    catagory_name.place(relx=0.218,rely=0.05)
    
    mainframe2=Ct.CTkFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=250,height=150,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.01,rely=0.08)
    
   
    enterAD_NO= Ct.CTkEntry(mainframe2, placeholder_text="Enter ",
        corner_radius=0,
            width=125,height=20
            )
                
    enterAD_NO.place(relx=0.45, rely=0.25)

    label1= Ct.CTkLabel(mainframe2,text="AD NO :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    label1.place(relx=0.1, rely=0.21)


    Submit1 = Ct.CTkButton(mainframe2,bg_color=('red'),text='SUBMIT',
                fg_color=('transparent')  ,width=5, height=2  ,
                font=("Stylus",20,'bold' ) ,
                command=lambda:Checking_Category("Classes 1-12.csv",enterAD_NO.get() ))
    Submit1.place(relx=0.35, rely=0.65)
    
    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback1 = Ct.CTkButton(mainframe,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:Goback1() )
    
    goback1.place(relx=0.0, rely=0.0)

    list1=['100MTS','200MTS','400MTS','800MTS','1500MTS(BOYS)','LONG JUMP','SHOT PUT','DISCUS THROW',
           'JAVALIN THROW','4 x RELAY']
    check_list=[]

    
    for i in list1:
        v=tk.IntVar()
        b=Ct.CTkCheckBox(mainframe1,text=i,font=('Stylus',20),text_color='white',variable=v)
        
        b.configure(command=lambda b=b,name_details=name_details,ad=ad,
                    v=v: tracer(b,name_details,ad,category,v))               
        check_list.append(b)
        var_list_junior.append(v)

    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
               column=i%n , sticky='W',pady=20 )
      
        mainframe1.grid_rowconfigure(i//n,weight=2)
        mainframe1.grid_columnconfigure(i%n,weight=2)
        
def senior(name_details,ad):                                                    #SENIOR FUNCTION
    global page5,root,mainframe
    page5.pack_forget()
    category='Senior'
    

    bg1 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic3 = Ct.CTkLabel(mainframe, image = bg1,text='') 
    pic3.place(relx=0, rely=0)

    mainframe1=Ct.CTkScrollableFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=800,height=490,
                    border_width=3,border_color='black')
    mainframe1.place(relx=0.218,rely=0.15)
    
    mainframe.place(relx=0,rely=0)
    
    student_detailSFrame(name_details,ad,category)
    
    catagory_name=Ct.CTkLabel(mainframe,fg_color='lime',bg_color='transparent',
                    width=820,height=50,text='SENIORS',
                    font=('Stylus',50,'bold'),text_color='white',                    
                    )
    catagory_name.place(relx=0.218,rely=0.05)

    mainframe2=Ct.CTkFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=250,height=150,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.01,rely=0.08)
    
   
    enterAD_NO= Ct.CTkEntry(mainframe2, placeholder_text="Enter ",
        corner_radius=0,
            width=125,height=20
            )
                
    enterAD_NO.place(relx=0.45, rely=0.25)

    label1= Ct.CTkLabel(mainframe2,text="AD NO :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    label1.place(relx=0.1, rely=0.21)


    Submit1 = Ct.CTkButton(mainframe2,bg_color=('red'),text='SUBMIT',
                fg_color=('transparent')  ,width=5, height=2  ,
                font=("Stylus",20,'bold' ) ,
                command=lambda:Checking_Category("Classes 1-12.csv",enterAD_NO.get() ))
    Submit1.place(relx=0.35, rely=0.65)
    
    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback1 = Ct.CTkButton(mainframe,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:Goback1() )
    
    goback1.place(relx=0.0, rely=0.0)

    list1=['100MTS','200MTS','400MTS','800MTS','1500MTS','LONG JUMP','SHOT PUT','DISCUS THROW',
           'JAVALIN THROW','4 x RELAY']
    
    for i in list1:
        v=tk.IntVar()
        b=Ct.CTkCheckBox(mainframe1,text=i,font=('Stylus',20),text_color='white',variable=v)
        
        b.configure(command=lambda b=b,name_details=name_details,ad=ad,
                    v=v: tracer(b,name_details,ad,category,v))
        check_list.append(b)
        var_list_senior.append(v)
    
    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
            column=i%n , sticky='W',pady=20 )
        mainframe1.grid_rowconfigure(i//n,weight=2)
        mainframe1.grid_columnconfigure(i%n,weight=2)

    

def tracer(checkbox,name_details,ad,category,v):
    global selected
    
    
    if len(selected)<3  and v.get()==1 :
        selected.append(checkbox.cget('text'))
        student_detailSFrame(name_details,ad,category,update=True)
    elif len(selected)<4 and v.get()==1 and (checkbox.cget('text') =='4 x RELAY' or '4 x RELAY' in selected) :
        selected.append(checkbox.cget('text'))
        student_detailSFrame(name_details,ad,category,update=True)
    elif v.get()==0:
        try:
            selected.remove(checkbox.cget('text'))
        except ValueError:
            pass
        finally:
            student_detailSFrame(name_details,ad,category,update=True)
    else:
        v.set(0)
        tk.messagebox.showerror("Error",'only 3 individual items and 1 group item can be selected')
     
def student_detailSFrame(name_details,ad,category,update=False):
    global mainframe,selected,data_file
         
    studentdetails=Ct.CTkFrame(mainframe,fg_color='#071952',bg_color='transparent',
                    width=250,height=400,
                    border_width=3,border_color='black')
    studentdetails.place(relx=0.01,rely=0.3)
    
    if update:
        event_label=Ct.CTkLabel(studentdetails,fg_color='transparent', bg_color='transparent',
                        text='',
                        font=('Stylus',20),text_color='white'                  
                        )
        
        event_label.place(relx=0.02,rely=0.4)
        event_label.configure(text=',\n'.join(selected))
        
    
    Adno_label=Ct.CTkLabel(studentdetails,fg_color='transparent',bg_color='transparent',
                        text='Adno  :',
                        font=('Stylus',20),text_color='white',                    
                        )
    Adno_label.place(relx=0.02,rely=0.1)
        
    ad_no=Ct.CTkLabel(studentdetails,fg_color='transparent',bg_color='transparent',
                        text=ad,
                        font=('Stylus',15),text_color='white',                    
                        )
    ad_no.place(relx=0.301,rely=0.1)

    Name_label=Ct.CTkLabel(studentdetails,fg_color='transparent',bg_color='transparent',
                        text='Name  :',
                        font=('Stylus',20),text_color='white',                    
                        )
    Name_label.place(relx=0.02,rely=0.2)
  
        
    name_no=Ct.CTkLabel(studentdetails,fg_color='transparent',bg_color='transparent',
                        text=name_details[0],
                        font=('Stylus',15),text_color='white',                    
                        )
    name_no.place(relx=0.305,rely=0.2)

    Class_label=Ct.CTkLabel(studentdetails,fg_color='transparent',bg_color='transparent',
                        text='Class  :',
                        font=('Stylus',20),text_color='white',                    
                        )
    Class_label.place(relx=0.02,rely=0.3)

    class_div=Ct.CTkLabel(studentdetails,fg_color='transparent',bg_color='transparent',
                        text=name_details[1],
                        font=('Stylus',15),text_color='white',                    
                        )
    class_div.place(relx=0.305,rely=0.3)  

    Submit_Button=Ct.CTkButton(studentdetails,bg_color=('red'),text='SUBMIT',
                fg_color=('transparent')  ,width=95, height=20  ,
                font=("Stylus",20,'bold' ) ,
                command=lambda:Insert("Classes 1-12.csv",ad,selected,category,data_file,Submit_Button )   )
    Submit_Button.place(relx=0.3, rely=0.85)
    Submit_Button.configure(state='normal')

def Checking_Category(c,adno):
    try:
        std=shelve.open('student_database.db',flag='r')
    except:
        f=open(c,'r')
        dict={}
        reader=csv.reader(f)
        std=shelve.open('student_database.db')
        for row in reader:
            dict[row[0]]=list(row[1:])
        std.update(dict)
        f.close()
    if adno.strip() in list(std.keys()):
        ad=adno.strip()
        name_details=std[ad]
        class_division=std[ad][1]
        j=class_division.split('-')   
        classofstudent=j[0]
        
        if True:
            tk.messagebox.showinfo('SUCESS','Student exists in the student database')
            check(classofstudent,name_details,ad)
            #d[adno]=std[adno]
            #check(std[adno])         
        else:
            tk.messagebox.showinfo('INFORMATION','Student should be in a class above IV.')          
    else:
        tk.messagebox.showerror("Error",'STUDENT DOES NOT EXIST')
   
def Insert(c,adno,selected,category,data_file,Submit_Button):
    Submit_Button.configure(state='disabled')
    try:
        std=shelve.open('student_database.db',flag='r')
    except:
        f=open(c,'r')
        dict={}
        reader=csv.reader(f)
        d=shelve.open(data_file)
        std=shelve.open('student_database.db')
        for row in reader:
            dict[row[0]]=list(row[1:])
        std.update(dict)
        f.close()
    d=shelve.open(data_file)
    if adno.strip() in list(std.keys()):
        for ad in std.keys():        
            if ad==adno:
                key=ad
                d[key]=std[ad]+[category]+[selected]
                tk.messagebox.showinfo('SUCESS','Items added sucessfully')
                break
    selected.clear()
    global var_list_kiddies,var_list_junior,var_list_subjunior,var_list_senior
    for i in [var_list_kiddies,var_list_junior,var_list_subjunior,var_list_senior]:
        for j in i:
            try:j.set(0)
            except:continue
        
    d.close()

def check(i,name_details,ad):    

    if i in ['I','II','III','IV']:
        return False
    elif i in ['V','VI']:
        kiddies(name_details,ad)
        return True
    elif i in['VII','VIII']:
        subjunior(name_details,ad)
        return True
    elif i in['IX','X']:
        junior(name_details,ad)
        return True
    elif i in['XI','XII']:
        senior(name_details,ad)
        return True

mainframe_2=Ct.CTkFrame(root,fg_color="#071952",
                          width=1366,height=720,border_color='black')
mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
def Match_Fixtures():
    global page5,root,mainframe_2
    page5.pack_forget()

    mainframe_2.place(relx=0,rely=0)
    mainframe_2.tkraise()

    bg2 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic2 = Ct.CTkLabel(mainframe_2, image = bg2,text='') 
    pic2.place(relx=0, rely=0)

    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=810,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)

    Kiddies = Ct.CTkButton(mainframe_2,text='KIDDIES',text_color='gold',
                       fg_color=('#071952'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),
                       width=150,height=50,     
                       command=lambda:kiddies_button() )

    Kiddies.place(relx=0.0, rely=0.38)


    SubJunior = Ct.CTkButton(mainframe_2,text='SUB JUNIOR',fg_color=('#071952'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=150,height=50,     
                       command=lambda:subjunior_button() )
    
    SubJunior.place(relx=0.0, rely=0.48)


    Junior = Ct.CTkButton(mainframe_2,text='JUNIOR',fg_color=('#071952'),
                       border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=150,height=50,     
                       command=lambda:junior_button() )
    
    Junior.place(relx=0.0, rely=0.58)

    Senior = Ct.CTkButton(mainframe_2,text='SENIOR',fg_color=('#071952'),
                           border_width=3,border_color='#687EFF',
                       font=("Plasma",20,'bold' ),text_color='gold',
                       width=150,height=50,     
                       command=lambda:senior_button() )
    
    Senior.place(relx=0, rely=0.68)
    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback2 = Ct.CTkButton(mainframe_2,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:GoBack2() )
    
    goback2.place(relx=0.0, rely=0.0)

def GoBack2():
    global page5,root,mainframe_2
    mainframe_2.place_forget()
    first_page()

Printframe=Ct.CTkFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=100,height=40,
                    border_width=3,border_color='black')
Kiddies_items=['50MTS','100MTS','STANDING BROAD JUMP','BALL THROW','4 x RELAY']
def kiddies_button():
    global mainframe2,mainframe
    Printframe.place_forget()
    mainframe2.pack_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)
    
    list1=['50MTS','100MTS','STANDING BROAD JUMP','BALL THROW','4 x RELAY']
    check_list=[]

    for i in list1:
        b=tk.Button(mainframe2,text=i,font=('Stylus',15,'bold'),bg=('#687EFF'),
                       fg=('white'))
        b.configure(command=lambda button_text=i: disp_Kiddies(button_text))
                 
        check_list.append(b)
    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
               column=i%n , sticky='W',pady=20 )
      
        mainframe2.grid_rowconfigure(i//n,weight=2)
        mainframe2.grid_columnconfigure(i%n,weight=2)
    
    Print_button = Ct.CTkButton(mainframe2,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Printcat(category='Kiddies'))
    Print_button.place(relx=0.45, rely=0.7)

SubJunior_items=['100MTS','200MTS','400MTS','LONG JUMP','SHOT PUT','4 x RELAY']
def subjunior_button():
    global mainframe2,mainframe
    Printframe.place_forget()
    mainframe2.pack_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)
    
    list1=['100MTS','200MTS','400MTS','LONG JUMP','SHOT PUT','4 x RELAY']
    check_list=[]

    for i in list1:
        b=tk.Button(mainframe2,text=i,font=('Stylus',15,'bold'),bg=('#687EFF'),
                       fg=('white'))
        b.configure(command=lambda button_text=i: disp_SubJunior(button_text))
                               
        check_list.append(b)
    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
               column=i%n , sticky='W',pady=20 )
      
        mainframe2.grid_rowconfigure(i//n,weight=2)
        mainframe2.grid_columnconfigure(i%n,weight=2)

    Print_button = Ct.CTkButton(mainframe2,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Printcat(category='Sub Junior'))
    Print_button.place(relx=0.55, rely=0.7)

Junior_items=['100MTS','200MTS','400MTS','800MTS','1500MTS(BOYS)','LONG JUMP','SHOT PUT','DISCUS THROW',
           'JAVALIN THROW','4 x RELAY']

def junior_button():
    global mainframe2,mainframe
    Printframe.place_forget()
    mainframe2.pack_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)

    list1=['100MTS','200MTS','400MTS','800MTS','1500MTS(BOYS)','LONG JUMP','SHOT PUT','DISCUS THROW',
           'JAVALIN THROW','4 x RELAY']
    check_list=[]

    for i in list1:
        b=tk.Button(mainframe2,text=i,font=('Stylus',15,'bold'),bg=('#687EFF'),
                       fg=('white'))
        b.configure(command=lambda button_text=i: disp_Junior(button_text))
                             
        check_list.append(b)
    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
               column=i%n , sticky='W',pady=20 )
      
        mainframe2.grid_rowconfigure(i//n,weight=2)
        mainframe2.grid_columnconfigure(i%n,weight=2)

    Print_button = Ct.CTkButton(mainframe2,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Printcat(category='Junior'))
    Print_button.place(relx=0.45, rely=0.9)

Senior_items=['100MTS','200MTS','400MTS','800MTS','1500MTS','LONG JUMP','SHOT PUT','DISCUS THROW',
           'JAVALIN THROW','4 x RELAY']
def senior_button():
    global mainframe2,mainframe
    Printframe.place_forget()
    mainframe2.pack_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)

    list1=['100MTS','200MTS','400MTS','800MTS','1500MTS','LONG JUMP','SHOT PUT','DISCUS THROW',
           'JAVALIN THROW','4 x RELAY']
    check_list=[]

    for i in list1:
        b=tk.Button(mainframe2,text=i,font=('Stylus',15,'bold'),bg=('#687EFF'),
                       fg=('white'))
        b.configure(command=lambda button_text=i: disp_Senior(button_text))
                            
        check_list.append(b)
    for i,j in enumerate(check_list):
        n=3
        j.grid(row=i//n,
               column=i%n , sticky='W',pady=20 )
      
        mainframe2.grid_rowconfigure(i//n,weight=2)
        mainframe2.grid_columnconfigure(i%n,weight=2)

    Print_button = Ct.CTkButton(mainframe2,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Printcat(category='Senior'))
    Print_button.place(relx=0.45, rely=0.9)

def Printcat(category):
    f=shelve.open('OUTDOOR.db')
    text=[]
    for i in f:  
        text.append([i]+f[i][:3])
    wb = openpyxl.Workbook() 
    sheet = wb.active
    a=1
    b=0
    for j in text:
        b=b+1
        a=1   
        if j[3]=='Senior':
            for i in j:            
                sheet.cell(row=b,column=a).value=i
                a=a+1
    wb.save(category+'.xlsx')
    tk.messagebox.showinfo("Sucess",' Excel file created')

def disp_Senior(text):
    global mainframe2,mainframe_2,Printframe
    Printframe.place_forget()
    mainframe2.pack_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)
    f=shelve.open(data_file)
    item=[]
    
    for i in f:
        item.append(f[i][3])
    l=[['Ad_No','NAME','CLASS/DIVISION',"CATEGORY","ITEMS"]]  

    for j in item:
        for k in j:    
            if k ==text:
                for i in f:                                
                   if f[i][2]=='Senior':
                        if text in f[i][3]:
                            if ([i]+f[i][:3]+[text]) in l:
                                pass
                            else:
                                l.append([i]+f[i][:3]+[text])
                            
    table = CTkTable(master=mainframe2, column=5,width=1,
                            values=l,anchor='w')
    table.pack(expand=True, fill="both", padx=20, pady=20)
    f.close()
    Printframe=Ct.CTkFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=100,height=40,
                    border_width=3,border_color='black')
    Printframe.place(relx=0.9,rely=0.3)

    Print_button = Ct.CTkButton(Printframe,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Print(text,l,category='Senior'))
    Print_button.place(relx=0, rely=0)

def disp_Junior(text):
    global mainframe2,mainframe_2,Printframe
    Printframe.place_forget()
    mainframe2.pack_forget()

    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)
    f=shelve.open(data_file)
    item=[]
    
    for i in f:
        item.append(f[i][3])
    l=[['Ad_No','NAME','CLASS/DIVISION',"CATEGORY","ITEMS"]]    
    for j in item:
        for k in j:    
            if k ==text:
                for i in f:                                
                   if f[i][2]=='Junior':
                        if text in f[i][3]:
                            if ([i]+f[i][:3]+[text]) in l:
                                pass
                            else:
                                l.append([i]+f[i][:3]+[text])
                        
    table = CTkTable(master=mainframe2, column=5,width=1,
                            values=l,anchor='w')
    table.pack(expand=True, fill="both", padx=20, pady=20)
    f.close()
    Printframe=Ct.CTkFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=100,height=40,
                    border_width=3,border_color='black')
    Printframe.place(relx=0.9,rely=0.3)

    Print_button = Ct.CTkButton(Printframe,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Print(text,l,category='Junior'))
    Print_button.place(relx=0, rely=0)

def disp_SubJunior(text):
    global mainframe2,mainframe_2,Printframe
    Printframe.place_forget()
    mainframe2.pack_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)
    f=shelve.open(data_file)
    item=[]
    
    for i in f:
        item.append(f[i][3])
    l=[['Ad_No','NAME','CLASS/DIVISION',"CATEGORY","ITEMS"]]    
    for j in item:
        for k in j:    
            if k ==text:
                for i in f:                                
                   if f[i][2]=='Sub Junior':
                        if text in f[i][3]:
                            if ([i]+f[i][:3]+[text]) in l:
                                pass
                            else:
                                l.append([i]+f[i][:3]+[text])
                        
    table = CTkTable(master=mainframe2, column=5,width=1,
                            values=l,anchor='w')
    table.pack(expand=True, fill="both", padx=20, pady=20)
    f.close()
    Printframe=Ct.CTkFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=100,height=40,
                    border_width=3,border_color='black')
    Printframe.place(relx=0.9,rely=0.3)

    Print_button = Ct.CTkButton(Printframe,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Print(text,l,category='Sub Junior'))
    Print_button.place(relx=0, rely=0)

def disp_Kiddies(text):
    global mainframe2,mainframe_2,Printframe
    Printframe.place_forget()
    mainframe2.place_forget()
    mainframe2=Ct.CTkScrollableFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
    mainframe2.place(relx=0.195,rely=0.15)
    f=shelve.open(data_file)
    item=[]
    
    for i in f:
        item.append(f[i][3])
    l=[['Ad_No','NAME','CLASS/DIVISION',"CATEGORY","ITEMS"]]    
    for j in item:
        for k in j:    
            if k ==text:
                for i in f:                                
                   if f[i][2]=='Kiddies':
                        if text in f[i][3]:
                            if ([i]+f[i][:3]+[text]) in l:
                                pass
                            else:
                                l.append([i]+f[i][:3]+[text])
                        
    table = CTkTable(master=mainframe2, column=5,width=1,
                            values=l,anchor='w')
    table.pack(expand=True, fill="both", padx=20, pady=20)
    f.close()
    Printframe=Ct.CTkFrame(mainframe_2,fg_color='#071952',bg_color='transparent',
                    width=100,height=40,
                    border_width=3,border_color='black')
    Printframe.place(relx=0.9,rely=0.3)

    Print_button = Ct.CTkButton(Printframe,bg_color=('transparent'),
                    text='PRINT',text_color='black',border_width=3,border_color='black',
                    fg_color=('white')  ,width=100, height=40  ,
                    command=lambda:Print(text,l,category='Kiddies'))
    Print_button.place(relx=0, rely=0)

def Print(item_name,text,category):
    wb = openpyxl.Workbook() 
    sheet = wb.active
    a=1
    b=0
    for i in text:
        b=b+1
        a=1
        for r in range(0,len(i)):
            sheet.cell(row=b,column=a).value=i[r]
            a=a+1
    wb.save(item_name+'-'+category+'.xlsx')
    tk.messagebox.showinfo("Sucess",' Excel file created')


mainframe3=Ct.CTkFrame(root,fg_color="#071952",
                          width=1366,height=720,border_color='black')
mainframe_3=Ct.CTkScrollableFrame(mainframe3,fg_color='#071952',bg_color='transparent',
                    width=850,height=490,
                    border_width=3,border_color='black')
mainframe_31=Ct.CTkScrollableFrame(mainframe_3,fg_color='#071952',bg_color='transparent',
                    width=750,height=450,
                    border_width=3,border_color='black')

def Result_Generation():
    global page5,root
    page5.pack_forget()
    
    mainframe3=Ct.CTkFrame(root,fg_color="#071952",
                          width=1366,height=720,border_color='black')
    mainframe3.place(relx=0,rely=0)

    bg1 = Ct.CTkImage( light_image=Image.open('5.jpg'),size=(1366,720) )
    pic1 = Ct.CTkLabel(mainframe3, image = bg1,text='') 
    pic1.place(relx=0, rely=0)

    mainframe_3=Ct.CTkFrame(mainframe3,fg_color='#071952',bg_color='transparent',
                    width=1250,height=535,
                    border_width=3,border_color='black')
    mainframe_3.place(relx=0.05,rely=0.13)

    Result_Generationlabel= Ct.CTkLabel(mainframe_3,text="RESULT GENERATION",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",30,'bold' )  )
    Result_Generationlabel.place(relx=0.05, rely=0.08)

    Winner= Ct.CTkEntry(mainframe_3, placeholder_text="Enter ",
        corner_radius=0,
            width=175,height=30
            )
    Winner.place(relx=0.2, rely=0.26)

    WinnerLabel= Ct.CTkLabel(mainframe_3,text="Winner        (AD NO) :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    WinnerLabel.place(relx=0.02, rely=0.25)

    RunnerUp1= Ct.CTkEntry(mainframe_3, placeholder_text="Enter ",
        corner_radius=0,
            width=175,height=30
            )
    RunnerUp1.place(relx=0.2, rely=0.36)

    RunnerUp1Label= Ct.CTkLabel(mainframe_3,text="RunnerUp 1 (AD NO) :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    RunnerUp1Label.place(relx=0.02, rely=0.35)

    RunnerUp2= Ct.CTkEntry(mainframe_3, placeholder_text="Enter ",
        corner_radius=0,
            width=175,height=30
            )
    RunnerUp2.place(relx=0.2, rely=0.46)

    RunnerUp2Label= Ct.CTkLabel(mainframe_3,text="RunnerUp 2 (AD NO) :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    RunnerUp2Label.place(relx=0.02, rely=0.45)

    Item= Ct.CTkEntry(mainframe_3, placeholder_text="Enter ",
        corner_radius=0,
            width=175,height=30
            )
    Item.place(relx=0.2, rely=0.56)
    Itemlabel= Ct.CTkLabel(mainframe_3,text="Item                           :",
            corner_radius=0,bg_color=('transparent'),
                fg_color=('transparent'),text_color='gold',
            font=("Stylus",20,'bold' )  )
    Itemlabel.place(relx=0.02, rely=0.55)

    Submit = Ct.CTkButton(mainframe_3,bg_color=('transparent'),
                        text='SUBMIT',text_color='gold',
                       fg_color=('black')  ,width=100, height=40  ,
                       border_width=3,border_color='black',
                       font=("Plasma",20,'bold' ),
                       command=lambda:Show_Result(mainframe_31,Winner.get(),RunnerUp1.get(),RunnerUp2.get(),Item.get()) )
    Submit.place(relx=0.15, rely=0.7)

    mainframe_31=Ct.CTkScrollableFrame(mainframe_3,fg_color='#071952',bg_color='transparent',
                    width=750,height=450,
                    border_width=3,border_color='black')
    mainframe_31.place(relx=0.36,rely=0.05)

    Print = Ct.CTkButton(mainframe_3,bg_color=('transparent'),
                        text='PRINT',text_color='gold',
                       fg_color=('black')  ,width=100, height=40  ,
                       border_width=3,border_color='black',
                       font=("Plasma",20,'bold' ),
                        )
    Print.place(relx=0.63, rely=0.92)

    bg6 = Ct.CTkImage( light_image=Image.open('6.png'),size=(100,40))
    goback3 = Ct.CTkButton(mainframe3,image=bg6,bg_color=('transparent'),
                           text='',text_color='black',
                       fg_color=('transparent')  ,width=100, height=40  ,
                       command=lambda:GoBack_3() )
    goback3.place(relx=0.0, rely=0.0)

def Show_Result(Frame,Winner,RunnerUp1,RunnerUp2,Item):
    
    f=shelve.open(data_file)
    Result=[['Ad_No','NAME','CLASS/DIVISION',"CATEGORY","ITEM"]]
    for i in f:
        if Winner in f:
            if Winner==i:
                Result.append([i]+f[i][:3]+[Item])
        else:
            tk.messagebox.showerror("Error",'Wrong Adno for Winner')
            break
        if RunnerUp1 in f:
            if RunnerUp1==i:
                Result.append([i]+f[i][:3]+[Item]) 
        else:
            tk.messagebox.showerror("Error",'Wrong Adno for RunnerUp 1')
            break
        if RunnerUp2 in f:
            if RunnerUp2==i:
                Result.append([i]+f[i][:3]+[Item])      
        else:
            tk.messagebox.showerror("Error",'Wrong Adno for RunnerUp 2')
            break
    table1 = CTkTable(master=Frame, column=5,width=1,
                            values=Result,anchor='w')
    table1.pack(expand=True,fill='both', padx=20, pady=20)
    f.close()
    

mainframe3=Ct.CTkFrame(root,fg_color="#071952",
                          width=1366,height=720,border_color='black')
def GoBack_3():  
    global mainframe3,root
    mainframe3.place_forget()
    first_page()  


root.mainloop()
